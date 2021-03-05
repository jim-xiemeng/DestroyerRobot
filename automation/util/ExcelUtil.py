#!/usr/bin/python3
# -*- coding: UTF-8 -*-
# @Time    : 2019/6/14 13:43
# @Author  : vivid
# @FileName: ExcelUtil.py
# @Software: PyCharm
# @email    ：331597811@QQ.com
import xlsxwriter
import xlrd
import re
from DestroyerRobot.automation.util.RandomUtil import TestRamdom
from DestroyerRobot.automation.util.SystemOsUtil import SystemOs
import openpyxl
class xlsxoper():
    """
    操作excel文件,文件写入即覆盖，不支持追加操作
    """
    def __init__(self,path):
        self.path = path
    def writeXLS(self,headings,data=None,datanums=20,sheet1='Sheet1'):
        """
        对文件是否存在进行判断,当文件不存在时自动创建文件
        :param headings:    #headings = ['Number', 'testA', 'testB']
        :param datanums:    默认循环次数，如果数据为None将循环20次
        :param sheet1:      页表
        :param data:
        #42-51行 ：判断是否采用随机数操作写入文件
        :return:
        """
        testos = SystemOs()
        if testos.is_file(self.path):
            try:
                workbook = xlsxwriter.Workbook(self.path)
                worksheet = workbook.add_worksheet(sheet1)
                # data = [
                #     ['2017-9-1', '2017-9-2', '2017-9-3', '2017-9-4', '2017-9-5', '2017-9-6'],
                #     [10, 40, 50, 20, 10, 50]
                # ]
                # worksheet.write_row('A1', headings)
                # worksheet.write_column('A2', data[0])
                # worksheet.write_column('B2', data[1])
                # worksheet.write_column('C2', data[2])
                if data is None:
                    data=[]
                    for j in range(len(headings)):
                        lists=[]
                        for i in range(datanums):
                            trandom = TestRamdom()
                            lists.append(trandom.RandomTest())
                        data.append(lists)
                else:
                    pass

                worksheet.write_row('A1', headings)
                column=[
                        'A2','B2','C2','D2','E2','F2',
                        'G2','H2','I2','J2','K2','L2',
                        'M2','N2','O2','P2','Q2','R2','S2',
                        'T2','U2','V2','W2','X2','Y2','Z2',
                        ]
                for i in range(len(headings)):
                    #print(data[i])
                    worksheet.write_column(column[i], data[i])
            except Exception as e:
                print(e)
            finally:
                workbook.close()
        else:
            msg = "文件不存在,创建新excel文件"
            import time
            time.sleep(5)
            try:
                # 新建一个excel文件
                workbook = xlsxwriter.Workbook(self.path)
                # 创建一个worksheet
                sheet = sheet1
                worksheet = workbook.add_worksheet(sheet)
                title = headings
                #向A1单元格写入title列表，列表中每一个字段对应一个内容
                worksheet.write_row(0,0,title)
                #关闭文件
                workbook.close()
                #重新写入文件
                datas = data
                datasnum = datanums
                self.writeXLS(title,datas,datasnum,sheet)
            except Exception as e:
                msg = "创建文件失败！"
            return msg

    def writeXLS_dict(self,addrows,columns,result,sheet1='Sheet1'):
        """
        以字典的形式进行写入，并保存到当前文件中row = 2 行数从第2行开始写入 , column = 1  列数从1开始
        :param addrows: excel中的行数
        :param columns: excel中的列数
        :param result:  写入excel中结果
        :param sheet1:  excel中页表
        :return:
        """
        testos = SystemOs()
        if testos.is_file(self.path):
            try:
                wb = openpyxl.load_workbook(self.path)
                ws = wb[sheet1]
                ws.cell(row = 1 +addrows,column=1+columns).value  = result
                wb.save(self.path)
            except Exception as e:
                print(e)
            finally:
                wb.close()


    def readerXLS(self,sheet1='Sheet1'):
        """
        将文件中的所有数据读取出来，放入list中，返回数据[['a','b'],['c','d']]
        :param sheet1: 页表
        :return:
        """
        testos = SystemOs()
        if testos.is_file(self.path):
            workbook = xlrd.open_workbook(self.path)
            booksheet = workbook.sheet_by_name(sheet1)
            # print(booksheet.nrows)
            # print(booksheet.ncols)
            p = list()
            for row in range(booksheet.nrows):
                row_data = []
                for col in range(booksheet.ncols):
                    cel = booksheet.cell(row, col)
                    val = cel.value
                    try:
                        val = cel.value
                        val = re.sub(r'\s+', '', val)
                    except:
                        pass
                    if type(val) == float:
                        val = int(val)
                    else:
                        val = str(val)
                    row_data.append(val)
                p.append(row_data)
            return p
        else:
            msg = "文件不存在"
            return msg

    def readerXLS_dict(self,sheet1='Sheet1'):
        """
        将文件中的所有数据读取出来，放入list中，读取出的结果以字典形式，前提是首行为字段标题
        :param sheet1: 页表
        :return:
        """
        testos = SystemOs()
        if testos.is_file(self.path):
            workbook = xlrd.open_workbook(self.path)
            booksheet = workbook.sheet_by_name(sheet1)
            # print(booksheet.nrows)
            # print(booksheet.ncols)
            p = list()
            for row in range(booksheet.nrows):
                row_data = []
                for col in range(booksheet.ncols):
                    cel = booksheet.cell(row, col)
                    val = cel.value
                    try:
                        val = cel.value
                        val = re.sub(r'\s+', '', val)
                    except:
                        pass
                    if type(val) == float:
                        val = int(val)
                    else:
                        val = str(val)
                    row_data.append(val)
                p.append(row_data)
                frist_p = p[0]
                last_p = p[1:]
                p_dict = []
                #for i in range(len(p) - 1):
                for i in range(last_p.__len__()):
                    x = dict(zip(frist_p, last_p[i]))
                    p_dict.append(x)
            return p_dict,frist_p
        else:
            msg = "文件不存在"
            return msg


    def cell_value(fp, sheet_name, row, col):
        """
        @Author  : ranshuang
        获取单元格的值，获取某一行某一列的值
        """
        test_data = xlrd.open_workbook(fp)
        sheet_name = test_data.sheet_by_name(sheet_name)
        return sheet_name.cell_value(row-1, col-1)


    def row_value(fp, sheet_name, case_name):
        """
        @Author  : ranshuang
        通过第一列信息内容，获取整行的值
        """
        test_data = xlrd.open_workbook(fp, 'br')
        sheet_name = test_data.sheet_by_name('%s' % sheet_name)
        cases = sheet_name.col_values(0)
        for case_i in range(len(cases)):
            if cases[case_i] == case_name:
                row_value = sheet_name.row_values(case_i)
                for i in range(0, row_value.__len__()):
                    if type(row_value[i]) == float:
                        row_value[i] = int(row_value[i])
                return row_value


    def col_value(fp, sheet_name, col):
        """
        @Author  : ranshuang
        指定列数，获取列信息，列数从1开始
        """
        test_data = xlrd.open_workbook(fp)
        sheet_name = test_data.sheet_by_name('%s' % sheet_name)
        return sheet_name.col_values(col - 1)



if __name__=='__main__':
    #DestroyerRobot/automation/datas_template/test_data.xlsx|C:\Users\vivid\PycharmProjects\untitled\DestroyerRobot\automation\datas_template\test_data.xlsx
    #file_readrec = 'C:\\Users\\vivid\\PycharmProjects\\untitled\\DestroyerRobot\\automation\\datas_template\\test_data.xlsx'
    file_addrec = 'C:\\Users\\vivid\PycharmProjects\\untitled\\DestroyerRobot\\automation\\api\data\\test_data.xlsx'
    #xls=xlsxoper(file_addrec)
    # p=xls.writeXLS('test_data')
    # print(p)
    # for i in p:
    #     print(i)
    #     for j in i:
    #         #print(j)
    #         pass

#    xls.do_excel()

    # data = [
    #     ['2017-9-1', '2017-9-2', '2017-9-3', '2017-9-4', '2017-9-5', '2017-9-6'],
    #     [10, 40, 50, 20, 10, 10000]
    # ]
    # p=xls.writeXLS(['序列号', '兑换码'],data)
    # print(p)
    #
    # xlsread=xlsxoper(file_addrec)
    # els=xlsread.readerXLS_dict("test_data")
    # print(els)
    # for i in els[0]:
    #     print(i)
    # print(els[1])
    # s = 'result_respon'
    # column1=els[1].index(s)
    #
    #
    #
    # # wb = openpyxl.load_workbook(file_addrec)
    # # ws = wb['test_data']
    # #
    # # # for i in range(els[0].__len__()):
    # # #     print(els[0][i])
    # # #     row = i+2
    # # #
    # # #     ws.cell(row,column=columns+1 ).value  = res
    # # ws.cell(row=2,column=columns+1 ).value  = res
    # #
    # # wb.save(file_addrec)
    # res="xiemeng"
    # for i in range(els[0].__len__()):
    #     rest = res+str(i)
    #     xlsread.writeXLS_dict(i,column1,rest,sheet1='test_data')

    file_path = "C:/Users/vivid/PycharmProjects/untitled/DestroyerRobot/automation/api/data/test_data.xlsx"
    column = 12
    caseNo = 1
    responseData ={'code': 200, 'msg': '成功', 'data': {'id': 90, 'code': None, 'userName': '13600000001', 'roleName': None, 'token': 'a258fd9f-9972-45d4-b2e1-8f913be314f2', 'loginName': '13600000001', 'loginSysName': 'bonus', 'menus': None, 'extend': {'id': 90, 'userName': '13600000001', 'jobNumber': '12344321', 'companyName': '优选好生活科技（珠海）有限公司', 'mobile': '13600000001', 'email': '1317@sina.com', 'pswd': '58323dbea8b88b453ff18d4aac3844c5', 'createTime': 1508314932000, 'lastLoginTime': 1599902037000, 'status': 1, 'cityId': 110100, 'roleId': 118, 'roleName': '运营人员', 'clientId': '60c5f16391d0cffe8de3e9f0846cf5ba'}, 'openId': None}}
    res=str(responseData)
    SheetName1 = "tapi"
    #addrows,columns,result,sheet1='Sheet1'
    xls = xlsxoper(file_path)
    xlss = xls.writeXLS_dict(caseNo,column,res,SheetName1)






