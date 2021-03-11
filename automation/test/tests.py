#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2020/8/15 11:48
# @Author : vivid
# @FileName: tests.py
# @Software: PyCharm
#encoding=utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import xlrd
from teddyAPI.utils.LogUtil import Log
logger = Log(logger='ParseExcel').get_log()

class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile =None
        self.font = Font(color = None) #设置字体颜色
        self.RGBDict = { 'red': 'FFFF3030', 'green': 'FF008B00' }

    def loadWorkBook(self,excelPathAndName):
        #将excel文件加载到内存，并获取workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as err:
            raise err
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        #根据sheet名获取该sheet对象
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as err:
            raise err

    def getSheetbyIndex(self,sheetIndex):
        #根据sheet的索引号获取该sheet对象
        try:
            sheetname = self.workbook.sheetnames[sheetIndex]
        except Exception as err:
            raise err
        sheet = self.workbook[sheetname]
        return  sheet

    def writeCell(self,sheet,content,rowNo = None,colsNo =None,style = None):
        #根据单元格excel中的编码坐标或者数字索引坐标向单元格中写入数据，
        #下标从1开始，参style表示字体的颜色的名字，比如red，green
        try:
            sheet.cell(row=rowNo,column=colsNo).value = content
            if style:
                sheet.cell(row=rowNo,column=colsNo).font = Font(color = self.RGBDict[style])
            self.workbook.save(self.excelFile)
        except Exception as e:
            raise e

    #按行读取excel数据，返回list类型
    def readExcel(self,fileName, SheetName="Sheet1"):
        data = xlrd.open_workbook(fileName)
        table = data.sheet_by_name(SheetName)
        # 获取总行数、总列数
        nrows = table.nrows
        if nrows > 1:
            # 获取第一行的内容，列表格式
            keys = table.row_values(0)
            # print(keys)
            listApiData = []
            # 获取每一行的内容，列表格式
            for col in range(1, nrows):
                values = table.row_values(col)
                logger.info("values: %s",values)
                # print(type(values))
                # keys，values这两个列表一一对应来组合转换为字典
                api_dict = dict(zip(keys, values))
                logger.info("api_dict:%s" % str(api_dict))
                # print(api_dict)
                listApiData.append(api_dict)
            return listApiData
        else:
            logger.info("表格未填写数据")
            return None

    #获取表格每列的index
    def trCol(self,fileName, SheetName="Sheet1",tr="trueResponse"):
        data = xlrd.open_workbook(fileName)
        table = data.sheet_by_name(SheetName)
        # 获取总行数、总列数
        nrows = table.nrows
        if nrows > 1:
            # 获取第一行的内容，列表格式
            keys = table.row_values(0)
            trcol = keys.index(tr)
            print(trcol)
            return trcol
        else:
            logger("表格未填写数据")
            return None

if __name__ == '__main__':
    #测试写入代码
    pe = ParseExcel()
    # pe.loadWorkBook(r'E:\pythonProject\\teddyAPI\TestData\\test_data.xlsx')
    # sheetObj = pe.getSheetByName(u"API")
    # print("通过名称获取sheet对象的名字：",sheetObj.title)
    # content = {"code": 200,"msg": "成功"}
    # wr = pe.writeCell(sheetObj,content,2,9)

    # s = ParseExcel().readExcel(file_path, "login")
    # print(s)
    file_path = "E:\pythonProject\\teddyAPI\TestData\\test_data.xlsx"
    tr = pe.trCol(file_path)
    print(tr)

    # column = pe.getColumn(sheetObj, 4)  # 获取第一行
    # for i in column:
    #     print(i.value)

    # print("通过index序号获取sheet对象的名字：",pe.getSheetbyIndex(0).title)
    # sheet = pe.getSheetbyIndex(0)
    # print(type(sheet))
    # print(pe.getRowsNumber(sheet)) #获取最大行号
    # print(pe.getColsNumber(sheet)) #获取最大列号
    # rows = pe.getRow(sheet,1)     #获取第一行
    # for i in rows:
    #     print(i.value)


